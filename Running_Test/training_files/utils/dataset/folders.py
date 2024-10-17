import torch.utils.data as data
import torch

from PIL import Image
import os
import scipy.io
import numpy as np
import csv
from openpyxl import load_workbook
import cv2

class LIVEC(data.Dataset):
    def __init__(self, root, index, transform):
        imgpath = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        imgpath = imgpath['AllImages_release']
        imgpath = imgpath[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, 'Images', imgpath[item][0][0]))
            gt.append(labels[item])
        gt = normalization(gt)

        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor

    def __len__(self):
        length = len(self.samples)
        return length


class Koniq10k(data.Dataset):
    def __init__(self, root, index, transform):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_distributions_sets.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, '1024x768', imgname[item]))
            gt.append(mos_all[item])
        gt = normalization(gt)

        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor

    def __len__(self):
        length = len(self.samples)
        return length


class SPAQ(data.Dataset):
    def __init__(self, root, index, transform):
        sample = []
        gt = []

        xls_file = os.path.join(root, 'Annotations', 'MOS_and_Image_attribute_scores.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        for count, row in enumerate(rows, 2):
            if count - 2 in index:
                sample.append(os.path.join(root, 'img', booksheet.cell(row=count, column=1).value))
                mos = booksheet.cell(row=count, column=2).value
                mos = np.array(mos)
                mos = mos.astype(np.float32)
                gt.append(mos)
            if count == 11126:
                break
        gt = normalization(gt)
        
        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor

    def __len__(self):
        length = len(self.samples)
        return length
    
class UHDIQA(data.Dataset):
    def __init__(self, root, index, transform):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'uhd-iqa-training-metadata.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['quality_mos'])).astype(np.float32)
                mos_all.append(mos)

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, 'challenge/training', imgname[item]))
            gt.append(mos_all[item])

        self.samples, self.gt = sample, gt
        self.transform = transform
        self.resize_size = [1907, 1231]

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform, self.resize_size)

        return img_tensor, gt_tensor

    def __len__(self):
        length = len(self.samples)
        return length

class BID(data.Dataset):
    def __init__(self, root, index, transform):

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = booksheet.cell(row=count, column=1).value
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = booksheet.cell(row=count, column=2).value
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break

        sample, gt = [], []
        for i, item in enumerate(index):
            sample.append(os.path.join(root, imgname[item]))
            gt.append(mos_all[item])
        gt = normalization(gt)

        self.samples, self.gt = sample, gt
        self.transform = transform

    def __getitem__(self, index):
        img_tensor, gt_tensor = get_item(self.samples, self.gt, index, self.transform)

        return img_tensor, gt_tensor

    def __len__(self):
        length = len(self.samples)
        return length

def get_item(samples, gt, index, transform, resize_size=[512, 512]):
    path, target = samples[index], gt[index]
    full_sample = load_image(path, resize_size)
    
    samples_full = transform(full_sample)

    return samples_full, torch.from_numpy(target).type(torch.FloatTensor)


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename

def load_image(img_path, resize_size):
    def resize(d_img, size):
        size = list(size)
        d_img = d_img.resize(size)
        return d_img
    
    d_img = Image.open(img_path).convert('RGB')
    full_img = resize(d_img, resize_size)
    
    return full_img

def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')
    
def normalization(data):
    data = np.array(data)
    range = np.max(data) - np.min(data)
    data = (data - np.min(data)) / range
    data = list(data.astype('float').reshape(-1, 1))

    return data